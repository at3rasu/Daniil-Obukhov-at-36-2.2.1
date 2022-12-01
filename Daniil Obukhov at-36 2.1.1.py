import csv
import os
from operator import itemgetter
from typing import List, Dict
import re
import pandas as pd
import openpyxl
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055
}


class Salary:
    def __init__(self, salary_from : str, salary_to : str, salary_currency : str):
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency
        self.average_salary = (int(float(salary_from) + float(salary_to)) / 2)


class Vacancy:
    def __init__(self, vacancy: Dict[str, str]):
        self.name = vacancy["name"]
        self.salary = Salary(salary_from=vacancy["salary_from"],
                             salary_to=vacancy["salary_to"],
                             salary_currency=vacancy["salary_currency"])
        self.area_name = vacancy["area_name"]
        self.published_at = vacancy["published_at"]
        self.year = self.published_at[:4]


class DataSet:
    def __init__(self, file_name : str):
        self.file_name = file_name
        self.vacancies_objects = self.__csv_reader()

    def __csv_reader(self) -> (List[Vacancy]):
        with open(self.file_name, encoding='utf-8-sig') as file:
            file_reader = csv.reader(file)
            lines = [row for row in file_reader]
            headlines, vacancies = lines[0], lines[1:]
        result = []
        for vacancy in vacancies:
            if (len(vacancy) == len(headlines)) and (all([v != "" for v in vacancy])):
                vacancy = [" ".join(re.sub("<.*?>", "", value).replace('\n', '; ').split()) for value in vacancy]
                vacancy = {x: y for x, y in zip([r for r in headlines], [v for v in vacancy])}
                vacancy = Vacancy(vacancy)
                result.append(vacancy)
        return result


class ParamSalary:
    def __init__(self, param : str, salary: Salary):
        self.param = param
        self.salary = int(salary.average_salary * currency_to_rub[salary.salary_currency])
        self.count_vacancy = 1

    def add_salary(self, new_salary : Salary) -> None:
        self.count_vacancy += 1
        self.salary = self.salary + new_salary.average_salary * currency_to_rub[new_salary.salary_currency]


def convert_to_param_salary(vacancies: List[Vacancy], comparison_param : str) -> (List[ParamSalary]):
    param_salary = {}
    for vacancy in vacancies:
        dict_comparison_params = {"year" : vacancy.year, "city" : vacancy.area_name}
        param = dict_comparison_params[comparison_param]
        if not param_salary.__contains__(param):
            param_salary[param] = ParamSalary(param, vacancy.salary)
        else:
            param_salary[param].add_salary(vacancy.salary)
    return [param_salary[d] for d in param_salary]


def convert_from_param_salary_to_dict(param_salary : List[ParamSalary]) -> (Dict[int, int], Dict[int, int]):
    return {x: y for x, y in zip([int(r.param) for r in param_salary], [0 if v.count_vacancy == 0 else int(v.salary / v.count_vacancy) for v in param_salary])}, {x: y for x, y in zip([int(r.param) for r in param_salary], [v.count_vacancy for v in param_salary])}


def add_missing_years(param_salary : List[ParamSalary]) -> List[ParamSalary]:
    years = [i.param for i in year_salary]
    s_years = [el.param for el in param_salary]
    for y in years:
        if y not in s_years:
            param_salary.insert(int(y) - int(years[0]), ParamSalary(y, Salary("0", "0", "RUR")))
            param_salary[int(y) - int(years[0])].count_vacancy = 0
    return param_salary


class Report:
    def __init__(self, profession : str, years: List[int], average_salary : List[int], average_salary_profession : List[int], count_vacancies_by_year : List[int], count_vacancies_by_year_prof : List[int], city_salary : Dict[str, int], city_vacancies : Dict[str, int], file_name):
        self.years = years
        self.average_salary = average_salary
        self.average_salary_profession = average_salary_profession
        self.count_vacancies_by_year = count_vacancies_by_year
        self.count_vacancies_by_year_prof = count_vacancies_by_year_prof
        self.city_salary = city_salary
        self.city_vacancies = city_vacancies
        self.profession = profession
        self.file_name = file_name

    def generate_excel(self) -> None:
        if not isinstance(self.file_name, str):
            raise TypeError('')
        if os.path.basename(self.file_name).split('.')[1] != "xlsx":
            raise TypeError('')
        if os.path.exists(self.file_name):
            raise FileExistsError("")
        columns = ["Год", "Средняя зарплата", f"Средняя зарплата - {self.profession}", "Количество вакансий", f"Количество вакансий - {self.profession}"]
        df = pd.DataFrame([[self.years[i], self.average_salary[i], self.average_salary_profession[i], self.count_vacancies_by_year[i], self.count_vacancies_by_year_prof[i]] for i in range(len(self.years))], columns=columns)
        cities_of_salary, salaries = [city for city in city_salary], [city_salary[city] for city in city_salary]
        cities_of_vacancy, vacancies = [city for city in city_vacancies], ['{:.2f}'.format(city_vacancies[city] * 100) + "%" for city in city_vacancies]
        df2 = pd.DataFrame([[cities_of_salary[i], salaries[i], "", cities_of_vacancy[i], vacancies[i]] for i in range(len(cities_of_salary))], columns=["Город", "Уровень зарплат", "", "Город", "Доля вакансий"])
        with pd.ExcelWriter(self.file_name, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name='Статистика по годам', index=False)
            df2.to_excel(writer, sheet_name="Статистика по городам", index=False)
        wb = openpyxl.load_workbook(self.file_name)
        worksheet1 = wb["Статистика по годам"]
        worksheet2 = wb["Статистика по городам"]
        thin = Side(border_style="thin")
        self.__add_border_and_align(worksheet1, thin, len(self.years) + 2, ["A", "B", "C", "D", "E"])
        self.__add_border_and_align(worksheet2, thin, max(len(cities_of_salary) + 2, len(cities_of_vacancy) + 2), ["A", "B", "C", "D", "E"])
        self.__make_max_column_width(worksheet1)
        self.__make_max_column_width(worksheet2)
        wb.save(self.file_name)

    def __add_border_and_align(self, worksheet : Worksheet, side : Side, count_columns : int, rows : List[str]) -> None:
        for i in range(1, count_columns):
            for row in rows:
                if i == 1:
                    worksheet[row + str(i)].alignment = Alignment(horizontal='left')
                if worksheet[row + str(i)].internal_value:
                    worksheet[row + str(i)].border = Border(top=side, bottom=side, left=side, right=side)

    def __make_max_column_width(self, worksheet : Worksheet) -> None:
        dims = {}
        for row in worksheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
                else:
                    dims[cell.column] = 0
        for col, value in dims.items():
            worksheet.column_dimensions[get_column_letter(col)].width = value + 2


input_data = []
for question in ["Введите название файла: ", "Введите название профессии: "]:
    print(question, end="")
    input_data.append(input())
data = DataSet(input_data[0]).vacancies_objects
data_profession = [d for d in data if input_data[1] in d.name]
year_salary = convert_to_param_salary(data, "year")
cities_salary = convert_to_param_salary(data, "city")
professions_year_salary = add_missing_years(convert_to_param_salary(data_profession, "year"))
city_salary = dict(sorted({x: y for x, y in zip([r.param for r in cities_salary], [int(v.salary / v.count_vacancy) for v in cities_salary])}.items(), key=itemgetter(1), reverse=True))
city_vacancies = dict(sorted({x: y for x, y in zip([r.param for r in cities_salary], [v.count_vacancy / len(data) for v in cities_salary])}.items(), key=itemgetter(1), reverse=True))
year_salary, year_vacancy = convert_from_param_salary_to_dict(year_salary)
professions_year_salary, professions_year_vacancies = convert_from_param_salary_to_dict(professions_year_salary)
city_salary = {x : y for x, y in zip([key for key in city_salary if city_vacancies[key] >= 0.01][:10], [city_salary[key] for key in city_salary if city_vacancies[key] >= 0.01])}
city_vacancies = {x : y for x, y in zip([key for key in city_vacancies if city_vacancies[key] >= 0.01][:10], [float('{:.4f}'.format(city_vacancies[key])) for key in city_vacancies if city_vacancies[key] >= 0.01])}
output_data = { "Динамика уровня зарплат по годам:" : year_salary,
                "Динамика количества вакансий по годам:" : year_vacancy,
                "Динамика уровня зарплат по годам для выбранной профессии:" : professions_year_salary,
                "Динамика количества вакансий по годам для выбранной профессии:" : professions_year_vacancies,
                "Уровень зарплат по городам (в порядке убывания):" : city_salary,
                "Доля вакансий по городам (в порядке убывания):" : city_vacancies}
[print(i, output_data[i]) for i in output_data]
report = Report(profession=input_data[1],
                years=[i for i in year_salary],
                average_salary=[year_salary[i] for i in year_salary],
                average_salary_profession=[professions_year_salary[i] for i in professions_year_salary],
                count_vacancies_by_year=[year_vacancy[i] for i in year_vacancy],
                count_vacancies_by_year_prof=[professions_year_vacancies[i] for i in professions_year_vacancies],
                city_salary=city_salary,
                city_vacancies=city_vacancies,
                file_name="report.xlsx")
report.generate_excel()
